"""
Whisper Accuracy Test with Medical Keyword Detection
Uses OpenAI's Whisper model to transcribe audio and match against medical keywords
"""

import gc
import json
import os
import random
import re
import sys
import threading
import time
from datetime import datetime

import numpy as np
import sounddevice as sd
import whisper
from flask import Flask, jsonify, render_template, send_file
from flask_socketio import SocketIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rapidfuzz import fuzz

# Configuration
WHISPER_MODEL = "base"  # OpenAI Whisper: tiny/base/small/medium/large
SAMPLE_RATE = 16000
RECORD_DURATION = 5  # seconds
CHUNK_SECONDS = 1.0  # Show progress every 1 second
CONFIDENCE_THRESHOLD = 80  # Fuzzy match threshold
MODEL_LOAD_TIMEOUT = 300  # 5 minutes timeout for model loading

# Get script directory for proper file paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Initialize Flask app
app = Flask(__name__, template_folder=".", static_folder="static")
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.jinja_env.auto_reload = True
socketio = SocketIO(app, cors_allowed_origins="*")


@app.after_request
def add_no_cache_headers(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


# Load medical supplies database
def load_medical_supplies():
    json_path = os.path.join(SCRIPT_DIR, "medical_supplies_expanded.json")
    with open(json_path, "r") as f:
        data = json.load(f)
    return {item["item"].lower(): item for item in data["medical_supplies"]}


# Initialize at module level with error handling
try:
    medical_supplies = load_medical_supplies()
    keywords = list(medical_supplies.keys())
except Exception as e:
    print(f"[ERROR] Failed to load medical supplies: {e}")
    medical_supplies = {}
    keywords = []

# Global state for UI
current_item = None
transcription_results = []

test_in_progress = False
test_lock = threading.Lock()
recorder = None
transcriber = None
matcher = None


class AudioRecorder:
    def __init__(self, sample_rate=SAMPLE_RATE, duration=RECORD_DURATION):
        self.sample_rate = sample_rate
        self.duration = duration
        self.audio_data = None

    def record(self):
        """Record audio from microphone and return float32 audio in [-1, 1]."""
        print(f"\n[MIC] Recording for {self.duration} seconds... (speak now)")
        try:
            self.audio_data = sd.rec(
                int(self.sample_rate * self.duration),
                samplerate=self.sample_rate,
                channels=1,
                dtype=np.int16,
            )
            sd.wait()
            print("[OK] Recording complete")
            audio = self.audio_data.astype(np.float32) / 32768.0
            return audio
        except Exception as e:
            print(f"[ERROR] Error recording audio: {e}")
            return None


class WhisperTranscriber:
    def __init__(self, model_name=WHISPER_MODEL):
        self.model_name = model_name
        self.model = None
        self.loading = False
        self.load_error = None
        self.load_event = threading.Event()
        self.load_thread = threading.Thread(target=self._load_model_async, daemon=True)
        self.load_thread.start()

    def _load_model_async(self):
        """Load model asynchronously in background thread."""
        try:
            self.loading = True
            print(
                f"[EXEC] Loading OpenAI Whisper {self.model_name} model in background..."
            )
            start_time = time.time()

            # Show progress dots
            def progress_indicator():
                dot_count = 0
                while self.loading and self.model is None:
                    print(".", end="", flush=True)
                    dot_count += 1
                    if dot_count >= 60:
                        elapsed = time.time() - start_time
                        print(f"\n[EXEC] Still loading... ({elapsed:.1f}s elapsed)")
                        dot_count = 0
                    time.sleep(0.5)

            progress_thread = threading.Thread(target=progress_indicator, daemon=True)
            progress_thread.start()

            # Load OpenAI Whisper model
            self.model = whisper.load_model(self.model_name)

            self.loading = False
            elapsed = time.time() - start_time
            print(f"\n[OK] Model loaded successfully ({elapsed:.1f}s)\n")
            self.load_event.set()

        except Exception as e:
            self.loading = False
            self.load_error = str(e)
            print(f"\n[ERROR] Model loading failed: {e}")
            print("[EXEC] Will retry on first transcription attempt")
            self.load_event.set()

    def wait_for_model(self, timeout=MODEL_LOAD_TIMEOUT):
        """Wait for model to load with timeout."""
        if self.model is not None:
            return True

        if not self.load_event.wait(timeout=timeout):
            raise TimeoutError(
                f"Model loading timed out after {timeout} seconds. "
                f"Check your internet connection and disk space."
            )

        if self.load_error:
            raise RuntimeError(f"Model loading failed: {self.load_error}")

        return self.model is not None

    def transcribe(self, audio_data):
        """Transcribe audio using OpenAI Whisper."""
        try:
            # Ensure model is loaded
            if self.model is None:
                if not self.wait_for_model():
                    return None

            audio = np.asarray(audio_data, dtype=np.float32)
            if audio.ndim == 2:
                audio = audio.squeeze()

            # Normalize audio
            max_val = np.max(np.abs(audio))
            if max_val > 0:
                audio = audio / (max_val + 1e-9)

            # Use OpenAI Whisper transcription
            result = self.model.transcribe(audio, language="en", verbose=False)
            text = result.get("text", "").strip()
            return {"text": text}

        except Exception as e:
            print(f"[ERROR] Transcription error: {type(e).__name__}: {e}")
            return None


class KeywordMatcher:
    def __init__(self, keywords, threshold=CONFIDENCE_THRESHOLD):
        self.keywords = keywords
        self.threshold = threshold
        self.normalized_keywords = [
            (keyword, re.sub(r"[^a-z0-9\s]", " ", keyword.lower()).strip())
            for keyword in keywords
        ]
        self.collapsed_keywords = [
            (keyword, normalized.replace(" ", ""))
            for keyword, normalized in self.normalized_keywords
        ]

    def find_matches(self, text):
        """Find matching keywords in transcribed text"""
        matches = []
        normalized_text = re.sub(r"[^a-z0-9\s]", " ", text.lower())
        collapsed_text = normalized_text.replace(" ", "")
        words = normalized_text.split()

        for word in words:
            for (keyword, normalized_keyword), (_, collapsed_keyword) in zip(
                self.normalized_keywords, self.collapsed_keywords
            ):
                ratio = fuzz.token_set_ratio(word, normalized_keyword)
                collapsed_ratio = fuzz.token_set_ratio(word, collapsed_keyword)

                if ratio >= self.threshold or collapsed_ratio >= self.threshold:
                    matches.append(
                        {"keyword": keyword, "word": word, "confidence": ratio}
                    )

        for (keyword, normalized_keyword), (_, collapsed_keyword) in zip(
            self.normalized_keywords, self.collapsed_keywords
        ):
            if len(normalized_keyword.split()) > 1:
                ratio = fuzz.token_set_ratio(normalized_text, normalized_keyword)
                collapsed_ratio = fuzz.token_set_ratio(
                    collapsed_text, collapsed_keyword
                )
                if ratio >= self.threshold:
                    if not any(m["keyword"] == keyword for m in matches):
                        matches.append(
                            {"keyword": keyword, "word": keyword, "confidence": ratio}
                        )
                elif collapsed_ratio >= self.threshold:
                    if not any(m["keyword"] == keyword for m in matches):
                        matches.append(
                            {
                                "keyword": keyword,
                                "word": keyword,
                                "confidence": collapsed_ratio,
                            }
                        )

        return sorted(matches, key=lambda x: x["confidence"], reverse=True)


@app.route("/")
def index():
    return render_template("test_ui.html")


@app.route("/api/supplies")
def get_supplies():
    """Get all medical supplies organized by rack"""
    supplies_by_rack = {}
    for item_name, item_data in medical_supplies.items():
        rack = item_data["rack"]
        if rack not in supplies_by_rack:
            supplies_by_rack[rack] = []
        supplies_by_rack[rack].append(
            {"name": item_name, "location": item_data["location"]}
        )
    return jsonify(supplies_by_rack)


@app.route("/api/model_info")
def get_model_info():
    """Get information about the speech recognition model being used"""
    return jsonify(
        {
            "model_type": "Whisper",
            "model_name": WHISPER_MODEL,
        }
    )


def run_single_test():
    """Run a single test cycle triggered by the UI."""
    global current_item, transcription_results, test_in_progress
    print("\n[EXEC] run_single_test() started")

    if not keywords:
        socketio.emit("test_error", {"message": "No keywords loaded."})
        return

    prompt_word = random.choice(keywords)
    socketio.emit("prompt_word", {"word": prompt_word})
    print(f"\n[TEST] Say this word: {prompt_word}")

    try:
        frames_per_chunk = max(1, int(SAMPLE_RATE * CHUNK_SECONDS))
        max_frames = int(SAMPLE_RATE * RECORD_DURATION)
        collected = []
        total_frames = 0
        transcription = ""
        matches = []

        print(f"[MIC] Recording for {RECORD_DURATION} seconds... (speak now)")
        socketio.emit("recording_started")
        print("[EXEC] Recording started, listening for audio...")

        # Record all audio
        with sd.InputStream(
            samplerate=SAMPLE_RATE,
            channels=1,
            dtype=np.int16,
        ) as stream:
            while total_frames < max_frames:
                try:
                    chunk, _ = stream.read(frames_per_chunk)
                    collected.append(chunk[:, 0])
                    total_frames += len(chunk)

                    # Show progress
                    progress_pct = (total_frames / max_frames) * 100
                    print(
                        f"[MIC] {progress_pct:.0f}% ({total_frames}/{max_frames} frames)"
                    )

                except Exception as chunk_err:
                    print(f"[ERROR] Error recording chunk: {chunk_err}")
                    continue

        print(f"[MIC] Recording complete. Total frames: {total_frames}")

        # Transcribe the complete audio once
        if collected:
            audio_data = np.concatenate(collected).astype(np.float32) / 32768.0
            print("[TEXT] Transcribing complete audio... (this may take a moment)")
            result = transcriber.transcribe(audio_data)

            if result is not None:
                transcription = result["text"].strip()
                print(f'[OK] Transcription: "{transcription}"')
                socketio.emit("realtime_transcription", {"text": transcription})
                matches = matcher.find_matches(transcription)
            else:
                print("[ERROR] Transcriber returned None")

        gc.collect()  # Free memory after recording
        print(
            f"[EXEC] Processing results... Transcription: '{transcription}', Matches: {len(matches)}"
        )

        if transcription == "":
            print("[EMIT] Sending test_error (no transcription)...")
            socketio.emit("test_error", {"message": "No transcription captured."})
            return

        if matches:
            print(f"\n[OK] Found {len(matches)} match(es):")
            top_match = matches[0]
            current_item = top_match["keyword"]
            for i, match in enumerate(matches, 1):
                keyword = match["keyword"]
                confidence = match["confidence"]
                location_data = medical_supplies[keyword]
                print(
                    f"  [{i}] '{match['word']}' → '{keyword}' ({confidence:.0f}% confidence)"
                )
                print(
                    f"       [LOC] Rack {location_data['rack']}, Location {location_data['location']}"
                )

            print("[EMIT] Sending item_detected event...")
            socketio.emit(
                "item_detected",
                {
                    "timestamp": datetime.now().isoformat(),
                    "transcription": transcription,
                    "matched_item": top_match["keyword"],
                    "confidence": top_match["confidence"],
                    "rack": medical_supplies[top_match["keyword"]]["rack"],
                    "location": medical_supplies[top_match["keyword"]]["location"],
                    "all_matches": matches,
                    "prompt_word": prompt_word,
                },
            )
            print("[EMIT] item_detected sent.")
        else:
            print("\n[ERROR] No matching keywords found")
            current_item = None
            print("[EMIT] Sending no_match event...")
            socketio.emit(
                "no_match",
                {
                    "timestamp": datetime.now().isoformat(),
                    "transcription": transcription,
                    "prompt_word": prompt_word,
                },
            )
            print("[EMIT] no_match sent.")

        print("[EXEC] Appending results to transcription_results...")
        transcription_results.append(
            {
                "test_number": len(transcription_results) + 1,
                "timestamp": datetime.now().isoformat(),
                "transcription": transcription,
                "matched_item": current_item,
                "matches_found": len(matches),
                "prompt_word": prompt_word,
                "model_type": "Whisper",
                "model_name": WHISPER_MODEL,
            }
        )
        print("[EXEC] Results appended. About to enter finally block...")
    finally:
        print("[EXEC] Entering finally block...")
        test_in_progress = False
        gc.collect()  # Free memory after test
        print("[EMIT] Sending test_complete event...")
        socketio.emit("test_complete")
        print("[EXEC] Test cycle complete.")


def print_summary():
    """Print test summary"""
    print(f"\n{'='*60}")
    print("TEST SUMMARY")
    print(f"{'='*60}")
    print(f"Total tests run: {len(transcription_results)}")

    successful_matches = sum(1 for r in transcription_results if r["matched_item"])
    print(f"Successful matches: {successful_matches}/{len(transcription_results)}")

    if successful_matches > 0:
        accuracy = (successful_matches / len(transcription_results)) * 100
        print(f"Match accuracy: {accuracy:.1f}%")

    results_path = os.path.join(SCRIPT_DIR, "test_results.json")
    with open(results_path, "w") as f:
        json.dump(transcription_results, f, indent=2)
    print(f"\nResults saved to: {results_path}")


def generate_excel_report():
    """Generate Excel report with test results and accuracy"""
    if not transcription_results:
        return None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"test_results_whisper_{timestamp}.xlsx"
    filepath = os.path.join(SCRIPT_DIR, filename)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"

    # Header styling
    header_fill = PatternFill(
        start_color="667EEA", end_color="667EEA", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = [
        "Test #",
        "Timestamp",
        "Prompt Word",
        "Transcription",
        "Matched Item",
        "Matches Found",
        "Success",
        "Model Type",
        "Model Name",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Data rows
    for idx, result in enumerate(transcription_results, 2):
        ws.cell(row=idx, column=1, value=result.get("test_number", idx - 1))
        ws.cell(row=idx, column=2, value=result.get("timestamp", ""))
        ws.cell(row=idx, column=3, value=result.get("prompt_word", ""))
        ws.cell(row=idx, column=4, value=result.get("transcription", ""))
        ws.cell(row=idx, column=5, value=result.get("matched_item", "No match"))
        ws.cell(row=idx, column=6, value=result.get("matches_found", 0))

        # Success column
        success = "Yes" if result.get("matched_item") else "No"
        success_cell = ws.cell(row=idx, column=7, value=success)
        if success == "Yes":
            success_cell.fill = PatternFill(
                start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"
            )
        else:
            success_cell.fill = PatternFill(
                start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"
            )

        ws.cell(row=idx, column=8, value=result.get("model_type", "Whisper"))
        ws.cell(row=idx, column=9, value=result.get("model_name", WHISPER_MODEL))

    # Calculate accuracy
    total_tests = len(transcription_results)
    successful_matches = sum(1 for r in transcription_results if r.get("matched_item"))
    accuracy = (successful_matches / total_tests * 100) if total_tests > 0 else 0

    # Add summary section
    summary_row = len(transcription_results) + 3
    ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=14)
    ws.cell(row=summary_row + 1, column=1, value="Total Tests:")
    ws.cell(row=summary_row + 1, column=2, value=total_tests)
    ws.cell(row=summary_row + 2, column=1, value="Successful Matches:")
    ws.cell(row=summary_row + 2, column=2, value=successful_matches)
    ws.cell(row=summary_row + 3, column=1, value="Accuracy:")
    accuracy_cell = ws.cell(row=summary_row + 3, column=2, value=f"{accuracy:.2f}%")
    accuracy_cell.font = Font(bold=True, size=12, color="667EEA")

    # Adjust column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 25

    # Save workbook
    wb.save(filepath)
    print(f"[OK] Excel report saved to: {filepath}")

    return filepath, accuracy


@socketio.on("start_test")
def handle_start_test():
    global test_in_progress
    with test_lock:
        if test_in_progress:
            return
        test_in_progress = True
    socketio.start_background_task(run_single_test)


@socketio.on("end_test")
def handle_end_test():
    """Handle end test request and generate Excel report"""
    global test_in_progress
    test_in_progress = False

    if not transcription_results:
        socketio.emit("test_error", {"message": "No test results to save."})
        return

    try:
        filepath, accuracy = generate_excel_report()
        total_tests = len(transcription_results)
        successful_matches = sum(
            1 for r in transcription_results if r.get("matched_item")
        )

        socketio.emit(
            "test_ended",
            {
                "total_tests": total_tests,
                "successful_matches": successful_matches,
                "accuracy": round(accuracy, 2),
                "filename": os.path.basename(filepath),
            },
        )

        print(
            f"\n[OK] Test ended. {successful_matches}/{total_tests} successful. Accuracy: {accuracy:.2f}%"
        )
    except Exception as e:
        print(f"[ERROR] Failed to generate report: {e}")
        socketio.emit("test_error", {"message": f"Failed to generate report: {str(e)}"})


@app.route("/download_report/<filename>")
def download_report(filename):
    """Download the generated Excel report"""
    filepath = os.path.join(SCRIPT_DIR, filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    return "File not found", 404


if __name__ == "__main__":
    try:
        print("=" * 60)
        print("WHISPER ACCURACY TEST - MEDICAL KEYWORD DETECTION")
        print("=" * 60)
        print("\nConfiguration:")
        print(f"  Model: OpenAI Whisper {WHISPER_MODEL}")
        print(f"  Sample Rate: {SAMPLE_RATE} Hz")
        print(f"  Recording Duration: {RECORD_DURATION} seconds")
        print(f"  Confidence Threshold: {CONFIDENCE_THRESHOLD}%")
        print(f"  Keywords Loaded: {len(keywords)}")

        recorder = AudioRecorder()
        transcriber = WhisperTranscriber()
        matcher = KeywordMatcher(keywords)

        print(
            "\n[EXEC] Waiting for Whisper model to load (this may take 30-60 seconds on first run)..."
        )
        transcriber.wait_for_model()

        print("\n[SETTINGS] Starting Flask web server...")
        print("[OK] Web UI started at: http://127.0.0.1:5000")
        print("[OK] Use the Start Test button in the UI")

        socketio.run(
            app,
            host="127.0.0.1",
            port=5000,
            debug=False,
            use_reloader=False,
            allow_unsafe_werkzeug=True,
        )

    except TimeoutError as e:
        print(f"\n[ERROR] {e}")
        print("[EXEC] Timeout waiting for model to load.")
        sys.exit(1)
    except KeyboardInterrupt:
        print("\n\n[STOP] Test interrupted by user")
        print_summary()
    except Exception as e:
        print(f"\n[ERROR] CRITICAL ERROR: {e}")
        import traceback

        traceback.print_exc()
        sys.exit(1)

    print("\n[OK] Test completed. Results saved to test_results.json")
